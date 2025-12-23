# excel_store.py
# ✅ 사용자가 "채택"했을 때만 엑셀(xlsx)에 저장합니다.

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

EXCEL_PATH = Path("adopted_shorts.xlsx")
SHEET_NAME = "adopted"

HEADERS = [
    "adopted_at",        # 채택 시간
    "video_length_sec",  # 영상 길이(초)
    "tone",              # 톤
    "character",         # 캐릭터(사람/오브젝트)
    "topic_keyword",     # 토픽 키워드
    "final_title",       # 최종 제목
    "scene_prompt",      # 한글 비디오 프롬프트(저장용)
    "script",            # 타임라인(저장용)
]

def _ensure_workbook() -> None:
    """엑셀 파일이 없으면 생성하고 헤더를 추가합니다."""
    if EXCEL_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(EXCEL_PATH)

def append_adopted_row(
    video_length_sec: int,
    tone: str,
    character: str,
    topic_keyword: str,
    final_title: str,
    scene_prompt: str,
    script: str,
) -> None:
    """채택된 결과 1건을 엑셀에 누적 저장합니다."""
    _ensure_workbook()

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        int(video_length_sec),
        tone.strip(),
        character.strip(),
        topic_keyword.strip(),
        final_title.strip(),
        scene_prompt.strip(),
        script.strip(),
    ])

    wb.save(EXCEL_PATH)
